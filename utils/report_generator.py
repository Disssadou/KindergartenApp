import logging
import io
from datetime import date
import calendar
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None
    get_column_letter = None
    PatternFill = None

logger = logging.getLogger("KindergartenApp.report_generator")

RUSSIAN_MONTHS_NOM = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}


def generate_attendance_excel_from_template(
    report_data: Dict[str, Any],
    staff_rates: Dict[int, float],
    default_rate: Optional[float],
    template_path: str,
    output_stream: io.BytesIO,
):
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required.")
    if not load_workbook or not get_column_letter or not PatternFill:
        raise ImportError("Essential openpyxl components could not be loaded.")

    FILL_WEEKEND_HOLIDAY = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )

    year = report_data.get("year", date.today().year)
    month = report_data.get("month", date.today().month)
    group_id_log = report_data.get("group_id", "N/A")
    logger.info(
        f"Populating report (template v11 - hide rows, no number_format) for Group ID {group_id_log}, {month}/{year}..."
    )

    try:
        wb = load_workbook(template_path)
        ws = wb.active
        if ws is None:
            raise ValueError("Could not get active sheet from template.")

        # --- 1. Заполнение Шапки (только динамические данные) ---
        month_name = RUSSIAN_MONTHS_NOM.get(month, "")
        ws["F4"] = f"{month_name} {year}"  # Месяц Год
        ws["D6"] = report_data.get("group_name")  # Название группы

        # --- 2. Данные для Таблицы ---
        children_data = report_data.get("children_data", [])
        days_in_month = report_data.get("days_in_month", 31)
        holiday_dates_set = {d.day for d in report_data.get("holiday_dates", [])}
        weekend_dates = set()
        for day_num in range(1, days_in_month + 1):
            try:
                d_obj = date(year, month, day_num)
            except ValueError:
                continue
            if d_obj.weekday() >= 5:
                weekend_dates.add(day_num)

        # --- Координаты ---

        header_day_digits_row_excel = 11
        data_start_row_excel = 14
        num_template_child_rows = 25
        total_row_excel = 40

        col_name_letter = "C"
        col_rate_letter = "E"
        col_day_start_letter = "F"

        col_total_present_letter = "AK"  # Всего посещено (+)
        col_total_absent_other_letter = "AL"  # "в т.ч. засчитываемых" (только 'н')
        col_payable_days_letter = "AM"  # Дни к оплате (+ и н)
        col_vacation_days_letter = "AN"  # Отпуск (о)
        col_sick_days_letter = "AO"  # Больничный (б)
        col_total_child_days_letter = "AP"  # "всего детодней" (рабочие дни месяца)

        # --- 3. Заполнение Данных Детей (только значения) ---
        num_children_filled = 0
        for i, child_data in enumerate(children_data):
            if i >= num_template_child_rows:
                logger.warning(
                    f"Template has only {num_template_child_rows} rows for children. Skipping child {i+1}."
                )
                break

            current_excel_row = data_start_row_excel + i
            num_children_filled += 1

            child_id = child_data.get("child_id")
            rate = staff_rates.get(child_id, default_rate)
            summary = child_data.get("summary", {})
            days_marks_data = child_data.get("days", {})

            ws[f"{col_name_letter}{current_excel_row}"].value = child_data.get(
                "child_name"
            )
            rate_cell = ws[f"{col_rate_letter}{current_excel_row}"]
            rate_cell.value = rate if rate is not None else ""
            if rate is not None:
                try:

                    calculated_monthly_rate = float(rate) * 22

                except (ValueError, TypeError):
                    logger.warning(
                        f"Could not convert rate '{rate}' to float for child {child_id}. Leaving rate cell empty."
                    )
                    calculated_monthly_rate = ""

            rate_cell.value = (
                calculated_monthly_rate if calculated_monthly_rate is not None else ""
            )

            # Дни 1-31
            for day_num in range(1, 32):
                current_col_idx = (
                    openpyxl.utils.column_index_from_string(col_day_start_letter)
                    + day_num
                    - 1
                )
                cell = ws.cell(row=current_excel_row, column=current_col_idx)
                mark_to_write = "х"

                if day_num <= days_in_month:
                    is_holiday = day_num in holiday_dates_set
                    is_weekend = day_num in weekend_dates
                    day_info = days_marks_data.get(str(day_num))

                    if is_holiday or is_weekend:
                        mark_to_write = ""
                    elif day_info:
                        api_mark = day_info.get("mark", "?")
                        mark_to_write = api_mark if api_mark is not None else "?"
                    else:
                        mark_to_write = "н"
                cell.value = str(mark_to_write)

            # Итоги по ребенку

            sick_days_val = summary.get("absent_sick_days", 0) or 0
            vacation_days_val = summary.get("absent_vacation_days", 0) or 0
            other_days_val = summary.get("absent_other_days", 0) or 0
            present_days_val = summary.get("present_days", 0) or 0
            payable_days_val = summary.get("payable_days", 0) or 0
            total_work_days_for_month = report_data.get("total_work_days", 0)

            # Рассчитываем "Пропущено дней (всего)"
            total_absent_val = sick_days_val + vacation_days_val + other_days_val

            # --- ОТЛАДКА ИТОГОВ для ребенка ---
            logger.debug(f"Child ID: {child_id}, Name: {child_data.get('child_name')}")
            logger.debug(f"  Present (+): {present_days_val}")
            logger.debug(f"  Sick (б): {sick_days_val}")
            logger.debug(f"  Vacation (о): {vacation_days_val}")
            logger.debug(f"  Other absent (н): {other_days_val}")
            logger.debug(f"  Calculated Total Absent (н+б+о): {total_absent_val}")
            logger.debug(f"  Payable days (+ и н): {payable_days_val}")
            logger.debug(
                f"  Total work days in month (для АP): {total_work_days_for_month}"
            )
            # ------------------------------------

            # Записываем итоги в колонки Excel согласно шаблону
            ws[f"AJ{current_excel_row}"].value = (
                present_days_val  # 36: Всего дней посещ. (+)
            )
            ws[f"AK{current_excel_row}"].value = (
                total_absent_val  # 37: Пропущено дней (всего)
            )
            ws[f"AL{current_excel_row}"].value = (
                other_days_val  # 38: в т.ч. засчитываемых (дни 'н')
            )
            ws[f"AM{current_excel_row}"].value = (
                payable_days_val  # 39: дни посещения, подлежащие оплате
            )
            ws[f"AN{current_excel_row}"].value = vacation_days_val  # 40: отпуск (о)
            ws[f"AO{current_excel_row}"].value = sick_days_val  # 41: больничный (б)
            ws[f"AP{current_excel_row}"].value = (
                total_work_days_for_month  # 42: всего детодней (рабочие дни месяца)
            )

        # --- 4. Заполнение Итоговой Строки "Всего детей" ---
        # Надпись "Всего детей" (C40:E40)
        daily_totals = report_data.get("daily_totals", {})
        for day_num in range(1, 32):
            current_col_idx = (
                openpyxl.utils.column_index_from_string(col_day_start_letter)
                + day_num
                - 1
            )
            cell = ws.cell(row=total_row_excel, column=current_col_idx)
            value_to_write_total = ""
            if day_num <= days_in_month:
                is_holiday = day_num in holiday_dates_set
                is_weekend = day_num in weekend_dates
                if not (is_holiday or is_weekend):
                    total_val = daily_totals.get(str(day_num), 0)
                    value_to_write_total = total_val if total_val is not None else 0
            cell.value = value_to_write_total

        # --- 5. ПРИМЕНЕНИЕ ЗАЛИВКИ для выходных/праздников (строки 11-40 Excel) ---
        start_fill_row_excel = (
            header_day_digits_row_excel  # Строка с цифрами дней "1", "2", ...
        )
        end_fill_row_excel = total_row_excel  # Итоговая строка "Всего детей"

        for day_num_fill in range(1, 32):
            col_idx_fill = (
                openpyxl.utils.column_index_from_string(col_day_start_letter)
                + day_num_fill
                - 1
            )
            apply_fill = False
            if day_num_fill <= days_in_month:
                is_holiday = day_num_fill in holiday_dates_set
                is_weekend = day_num_fill in weekend_dates
                if is_holiday or is_weekend:
                    apply_fill = True

            if apply_fill and FILL_WEEKEND_HOLIDAY:
                for r_fill_idx in range(start_fill_row_excel, end_fill_row_excel + 1):
                    cell_to_fill = ws.cell(row=r_fill_idx, column=col_idx_fill)

                    cell_to_fill.fill = FILL_WEEKEND_HOLIDAY

        # --- 6. СКРЫТИЕ ЛИШНИХ СТРОК ДЛЯ ДАННЫХ ДЕТЕЙ ---
        first_row_to_hide_excel = data_start_row_excel + num_children_filled
        last_template_data_row_excel = (
            data_start_row_excel + num_template_child_rows - 1
        )

        end_row_for_hiding_excel = last_template_data_row_excel

        if first_row_to_hide_excel <= end_row_for_hiding_excel:
            logger.info(
                f"Hiding unused child data rows from Excel row {first_row_to_hide_excel} to {end_row_for_hiding_excel}"
            )
            for r_idx in range(first_row_to_hide_excel, end_row_for_hiding_excel + 1):

                if r_idx in ws.row_dimensions:
                    ws.row_dimensions[r_idx].hidden = True
                else:

                    ws.row_dimensions[r_idx].hidden = True

        # --- 7. Заполнение Подвала (только ФИО воспитателя) ---
        initial_footer_teacher_row_excel = 46
        current_footer_teacher_row_excel = initial_footer_teacher_row_excel

        num_rows_deleted_or_hidden = 0
        if first_row_to_hide_excel <= last_template_data_row_excel:
            num_rows_deleted_or_hidden = (
                last_template_data_row_excel - first_row_to_hide_excel + 1
            )

        ws[f"H{initial_footer_teacher_row_excel}"] = report_data.get("teacher_name", "")

        # --- Сохранение в поток ---
        wb.save(output_stream)
        logger.info(
            "Attendance report Excel file (template v11 - hide rows, minimal style) generated successfully."
        )

    except Exception as e:
        logger.exception("Failed to generate attendance Excel report.")
        raise
